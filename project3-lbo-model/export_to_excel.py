"""
export_to_excel.py
------------------
Professional multi-sheet LBO workbook — formula-driven where possible.

Formula strategy
----------------
  Income Statement  : fully formula-driven from an editable driver block
                      (change revenue growth or margins → P&L recalculates)
  Cash Flow Stmt    : Net Income & D&A cross-reference IS; CapEx formula from
                      IS Revenue; Operating CF / FCF / Net Change as formulas
  Balance Sheet     : item values from Python model; totals & BS Check as formulas
  Sources & Uses    : total rows are SUM formulas
  Debt Schedule     : Ending Balance = Begin − Req Amort − Cash Sweep (formula)
  Returns Bridge    : Total Value Creation and Exit Equity as formulas
  Interest Expense  : hardcoded (computed by iterative debt engine; not replicable
                      in a simple Excel formula — noted with ‡ footnote)

Usage
-----
    from export_to_excel import build_excel_workbook
    xl_bytes = build_excel_workbook(assumptions)   # → bytes

    # Standalone: python export_to_excel.py → DollarGeneral_LBO_Model.xlsx
"""

import io
import numpy as np
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from model.assumptions import DealAssumptions, base_case
from model.lbo_engine import run_model
from model.debt_schedule import sources_uses_df
from analysis.scenarios import run_scenarios
from analysis.credit_metrics import build_credit_dashboard


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
COLOR_NAVY_DARK  = "1F2D40"
COLOR_NAVY_MID   = "2E4057"
COLOR_NAVY_INPUT = "1A4A7A"   # slightly lighter navy for editable driver block
COLOR_GOLD       = "C9A84C"
COLOR_GOLD_LIGHT = "F5E6C3"
COLOR_ROW_ALT    = "EEF2F7"
COLOR_WHITE      = "FFFFFF"
COLOR_LIGHT_GRAY = "F7F9FC"
COLOR_SECTION_BG = "E8EDF3"
COLOR_MID_GRAY   = "D6DCE4"
COLOR_DARK_TEXT  = "1A1A2E"
COLOR_GREEN      = "1A6B3C"
COLOR_RED_DARK   = "8B1A1A"
COLOR_INPUT_BG   = "EBF3FF"   # light blue tint for editable driver cells


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=COLOR_DARK_TEXT, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _border(style="thin"):
    s = Side(style=style, color="B0BAC8")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, row, col, value, bold=False, font_size=10,
           fg=None, font_color=COLOR_DARK_TEXT,
           align="left", wrap=False, italic=False,
           num_format=None, border=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=font_size, color=font_color, italic=italic)
    cell.alignment = _align(h=align, wrap=wrap)
    if fg:
        cell.fill = _fill(fg)
    if num_format:
        cell.number_format = num_format
    if border:
        cell.border = _border()
    return cell

def _col_header(ws, row, col, label, bg=COLOR_NAVY_DARK):
    _write(ws, row, col, label,
           bold=True, font_size=10, fg=bg, font_color=COLOR_WHITE,
           align="center", border=True)

def _section_header(ws, row, col, label, ncols=1, bg=COLOR_NAVY_MID):
    _write(ws, row, col, label,
           bold=True, font_size=11, fg=bg, font_color=COLOR_WHITE,
           align="left", border=True)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + ncols - 1)

def _title_row(ws, row, label, ncols, last_col_idx):
    ws.merge_cells(start_row=row, start_column=2,
                   end_row=row,   end_column=last_col_idx)
    cell = ws.cell(row=row, column=2, value=label)
    cell.font      = _font(bold=True, size=14, color=COLOR_WHITE)
    cell.fill      = _fill(COLOR_NAVY_DARK)
    cell.alignment = _align(h="center", v="center")
    ws.row_dimensions[row].height = 22


# ---------------------------------------------------------------------------
# Column-letter helper
# ---------------------------------------------------------------------------

def _yc(yr: int) -> str:
    """Column letter for year yr (1-based). Year 1 → 'C', Year 2 → 'D', …"""
    return get_column_letter(2 + yr)


# ---------------------------------------------------------------------------
# Generic value/formula cell writer
# ---------------------------------------------------------------------------

def _vcell(ws, row, col, val_or_formula, bold=False, bg=COLOR_WHITE,
           fc=COLOR_DARK_TEXT, pct=False, mult=False, border=True):
    """Write a data cell (value or formula string)."""
    cell = ws.cell(row=row, column=col, value=val_or_formula)
    cell.font      = _font(bold=bold, color=fc)
    cell.fill      = _fill(bg)
    cell.alignment = _align(h="right")
    if border:
        cell.border = _border()
    if pct:
        cell.number_format = "0.0%"
    elif mult:
        cell.number_format = '0.00"x"'
    else:
        cell.number_format = "#,##0.0"
    return cell


# ===========================================================================
# Sheet 1 — Cover
# ===========================================================================

def _build_cover(wb, assumptions, result):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22

    r = result["returns"]
    a = assumptions

    ws.merge_cells("B2:F4")
    cell = ws["B2"]
    cell.value     = "LEVERAGED BUYOUT ANALYSIS\nDollar General Corporation (NYSE: DG)"
    cell.font      = _font(bold=True, size=18, color=COLOR_WHITE)
    cell.fill      = _fill(COLOR_NAVY_DARK)
    cell.alignment = _align(h="center", v="center", wrap=True)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 20

    ws.merge_cells("B5:F5")
    cell = ws["B5"]
    cell.value     = (f"Discount Retail  |  ~20,000 Stores  |  "
                      f"Prepared: {datetime.today().strftime('%B %d, %Y')}")
    cell.font      = _font(italic=True, size=10, color=COLOR_GOLD_LIGHT if False else "CCCCCC")
    cell.fill      = _fill(COLOR_NAVY_DARK)
    cell.alignment = _align(h="center", v="center")
    ws.row_dimensions[5].height = 16

    row = 7
    _section_header(ws, row, 2, "KEY RETURNS SUMMARY", ncols=5)
    row += 1

    irr_val  = r["irr"]
    moic_val = r["moic"]

    metrics = [
        ("IRR",                f"{irr_val:.1%}" if not np.isnan(irr_val) else "N/A"),
        ("MOIC",               f"{moic_val:.2f}x"),
        ("Entry EV ($M)",      f"${r['entry_ev']:,.0f}"),
        ("Equity Invested ($M)", f"${r['entry_equity']:,.0f}"),
        ("Exit EV ($M)",       f"${r['exit_ev']:,.0f}"),
        ("Exit Equity ($M)",   f"${r['exit_equity']:,.0f}"),
        ("Hold Period",        f"{a.hold_years} years"),
        ("Entry EV/EBITDA",    f"{a.entry_ev_multiple:.1f}x"),
        ("Exit EV/EBITDA",     f"{a.exit_ev_multiple:.1f}x"),
        ("Entry Leverage",     f"{a.entry_leverage:.1f}x"),
    ]

    for i, (lbl, val) in enumerate(metrics):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, row, 2, lbl, bold=True, fg=bg, border=True)
        _write(ws, row, 3, val, fg=bg, align="right", border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4).fill = _fill(bg)
        row += 1

    row += 1
    _section_header(ws, row, 2, "TRANSACTION OVERVIEW", ncols=5)
    row += 1

    overview = [
        ("Target",              "Dollar General Corporation"),
        ("Sector",              "Consumer Discretionary — Discount Retail"),
        ("LTM Revenue ($M)",    f"${a.entry_revenue:,.0f}"),
        ("LTM EBITDA ($M)",     f"${a.entry_ebitda:,.0f}"),
        ("LTM EBITDA Margin",   f"{a.entry_ebitda / a.entry_revenue:.1%}"),
        ("Total New Debt ($M)", f"${a.total_new_debt:,.0f}"),
        ("Sponsor Equity ($M)", f"${a.sponsor_equity:,.0f}"),
        ("Equity / Total Cap.", f"{a.sponsor_equity / (a.total_new_debt + a.sponsor_equity):.1%}"),
    ]

    for i, (lbl, val) in enumerate(overview):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, row, 2, lbl, bold=True, fg=bg, border=True)
        _write(ws, row, 3, val, fg=bg, align="right", border=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4).fill = _fill(bg)
        row += 1

    row += 2
    ws.merge_cells(f"B{row}:F{row}")
    cell = ws[f"B{row}"]
    cell.value = ("CONFIDENTIAL — For Discussion Purposes Only. "
                  "All values USD millions unless noted. "
                  "‡ Interest Expense is computed by an iterative debt engine and is a fixed input to the P&L.")
    cell.font      = _font(size=8, italic=True, color="888888")
    cell.alignment = _align(h="center", wrap=True)
    ws.row_dimensions[row].height = 28


# ===========================================================================
# Sheet 2 — Assumptions
# ===========================================================================

def _build_assumptions_sheet(wb, assumptions):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 20
    for col in "DEFGHI":
        ws.column_dimensions[col].width = 16

    a = assumptions
    n = a.hold_years

    _title_row(ws, 1, "DEAL ASSUMPTIONS", n, 2 + n)

    r = 3
    _section_header(ws, r, 2, "1.  ENTRY / TRANSACTION", ncols=4)
    r += 1

    entry_items = [
        ("LTM EBITDA ($M)",          a.entry_ebitda,        '#,##0.0'),
        ("Entry EV/EBITDA (x)",       a.entry_ev_multiple,   '0.0"x"'),
        ("Entry EV ($M)",             a.entry_ev,            '#,##0.0'),
        ("Existing Net Debt ($M)",    a.existing_net_debt,   '#,##0.0'),
        ("Total New Debt ($M)",       a.total_new_debt,      '#,##0.0'),
        ("Sponsor Equity ($M)",       a.sponsor_equity,      '#,##0.0'),
        ("Entry Gross Leverage (x)",  a.entry_leverage,      '0.00"x"'),
        ("Hold Period (years)",       a.hold_years,          '0'),
        ("Exit EV/EBITDA (x)",        a.exit_ev_multiple,    '0.0"x"'),
        ("Tax Rate",                  a.tax_rate,            '0.0%'),
        ("Cash Sweep (%)",            a.cash_sweep_pct,      '0%'),
        ("Min Cash Balance ($M)",     a.min_cash_balance,    '#,##0.0'),
    ]

    for i, (lbl, val, fmt) in enumerate(entry_items):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, r, 2, lbl, bold=False, fg=bg, align="left", border=True)
        cell = ws.cell(row=r, column=3, value=val)
        cell.font = _font(); cell.fill = _fill(bg)
        cell.alignment = _align(h="right"); cell.border = _border()
        cell.number_format = fmt
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4).fill = _fill(bg)
        r += 1

    r += 1
    _section_header(ws, r, 2, "2.  DEBT STRUCTURE", ncols=7)
    r += 1
    for i, h in enumerate(["Tranche", "Principal ($M)", "Rate Type", "All-In Rate",
                            "Term (yrs)", "Amort %", "OID %"]):
        _col_header(ws, r, 2 + i, h)
    r += 1

    for j, t in enumerate(a.debt_tranches):
        bg = COLOR_ROW_ALT if j % 2 else COLOR_WHITE
        rate_type = f"SOFR + {t.spread:.0%}" if t.is_floating else "Fixed"
        row_vals  = [t.name, t.principal, rate_type, t.all_in_rate,
                     t.term_years, t.amort_pct, t.oid_pct]
        row_fmts  = [None, '#,##0.0', None, '0.00%', '0', '0.0%', '0.00%']
        for k, (v, f) in enumerate(zip(row_vals, row_fmts)):
            cell = ws.cell(row=r, column=2 + k, value=v)
            cell.font = _font(); cell.fill = _fill(bg)
            cell.alignment = _align(h="right" if k > 0 else "left")
            cell.border = _border()
            if f: cell.number_format = f
        r += 1

    r += 1
    _section_header(ws, r, 2, "3.  OPERATING PROJECTIONS (BY YEAR)", ncols=2 + n)
    r += 1
    _write(ws, r, 2, "Driver", bold=True, fg=COLOR_SECTION_BG, border=True)
    for k in range(n):
        _col_header(ws, r, 3 + k, f"Year {k+1}", bg=COLOR_SECTION_BG)
        ws.cell(row=r, column=3+k).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    op_items = [
        ("Revenue Growth (%)",  a.revenue_growth,    '0.0%'),
        ("EBITDA Margin (%)",   a.ebitda_margin,     '0.0%'),
        ("D&A (% Revenue)",     a.da_pct_revenue,    '0.0%'),
        ("CapEx (% Revenue)",   a.capex_pct_revenue, '0.0%'),
    ]
    for i, (lbl, vals, fmt) in enumerate(op_items):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, r, 2, lbl, bold=False, fg=bg, border=True)
        for k, v in enumerate(vals):
            cell = ws.cell(row=r, column=3 + k, value=v)
            cell.font = _font(); cell.fill = _fill(bg)
            cell.alignment = _align(h="right"); cell.border = _border()
            cell.number_format = fmt
        r += 1


# ===========================================================================
# Sheet 3 — Sources & Uses  (SUM formula totals)
# ===========================================================================

def _build_sources_uses(wb, assumptions):
    ws = wb.create_sheet("Sources & Uses")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 4
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12

    _title_row(ws, 1, "SOURCES & USES OF FUNDS", 1, 8)

    sources_df, uses_df = sources_uses_df(assumptions)

    r = 3
    for col, lbl, c in [("B", "SOURCES", 2), ("C", "Amount ($M)", 3), ("D", "% Total", 4),
                         ("F", "USES",    6), ("G", "Amount ($M)", 7), ("H", "% Total", 8)]:
        _col_header(ws, r, c, lbl)
    r += 1

    # Track amount cells for SUM formula
    src_start = r
    uses_start = r

    max_rows = max(len(sources_df), len(uses_df))
    for i in range(max_rows):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        is_last_src  = (i == len(sources_df) - 1)
        is_last_uses = (i == len(uses_df) - 1)

        # Sources side
        if i < len(sources_df):
            row_s  = sources_df.iloc[i]
            is_tot = "Total" in str(row_s["Item"])
            bold   = is_tot
            fg_s   = COLOR_SECTION_BG if is_tot else bg
            _write(ws, r, 2, row_s["Item"], bold=bold, fg=fg_s, border=True)
            if is_tot:
                # SUM formula
                sum_f = f"=SUM(C{src_start}:C{r-1})"
                cell  = ws.cell(row=r, column=3, value=sum_f)
            else:
                cell  = ws.cell(row=r, column=3, value=row_s["Amount ($M)"])
            cell.font = _font(bold=bold); cell.fill = _fill(fg_s)
            cell.alignment = _align(h="right"); cell.border = _border()
            cell.number_format = '#,##0.0'
            # % of total stays as text (pre-formatted)
            _write(ws, r, 4, row_s["% of Total"], bold=bold, fg=fg_s, align="right", border=True)
        else:
            for c in [2, 3, 4]:
                ws.cell(row=r, column=c).fill = _fill(bg)

        ws.cell(row=r, column=5).fill = _fill(COLOR_WHITE)

        # Uses side
        if i < len(uses_df):
            row_u  = uses_df.iloc[i]
            is_tot = "Total" in str(row_u["Item"])
            bold   = is_tot
            fg_u   = COLOR_SECTION_BG if is_tot else bg
            _write(ws, r, 6, row_u["Item"], bold=bold, fg=fg_u, border=True)
            if is_tot:
                sum_f = f"=SUM(G{uses_start}:G{r-1})"
                cell  = ws.cell(row=r, column=7, value=sum_f)
            else:
                cell  = ws.cell(row=r, column=7, value=row_u["Amount ($M)"])
            cell.font = _font(bold=bold); cell.fill = _fill(fg_u)
            cell.alignment = _align(h="right"); cell.border = _border()
            cell.number_format = '#,##0.0'
            _write(ws, r, 8, row_u["% of Total"], bold=bold, fg=fg_u, align="right", border=True)
        else:
            for c in [6, 7, 8]:
                ws.cell(row=r, column=c).fill = _fill(bg)

        r += 1

    r += 2
    ws.merge_cells(f"B{r}:H{r}")
    a = assumptions
    cell = ws[f"B{r}"]
    cell.value = (f"Entry Leverage: {a.entry_leverage:.1f}x  |  "
                  f"Net Leverage: {a.entry_net_leverage:.1f}x  |  "
                  f"Equity / Total Cap: {a.sponsor_equity/(a.total_new_debt+a.sponsor_equity):.1%}")
    cell.font = _font(size=9, italic=True, color="555555")
    cell.alignment = _align(h="center")


# ===========================================================================
# Sheet 4 — Income Statement  (FULLY FORMULA-DRIVEN)
# ===========================================================================

def _build_income_statement(wb, is_df, assumptions) -> dict:
    """
    Builds a formula-driven IS with an editable driver block.
    Returns row_map {line_item: {yr: cell_address}} for cross-sheet references.
    """
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    n = assumptions.hold_years
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    # ── Pre-defined row layout ──────────────────────────────────────────────
    R_TITLE      = 1
    R_HDRS       = 3
    R_DRV_HDR    = 4   # "REVENUE & MARGIN DRIVERS" section header
    R_ENTRY_REV  = 5   # Entry LTM Revenue  — input value
    R_GROWTH     = 6   # Revenue Growth %   — input per year
    R_EBITDA_MRG = 7   # EBITDA Margin %    — input per year
    R_DA_PCT     = 8   # D&A % Revenue      — input per year
    R_GM_PCT     = 9   # Gross Margin %     — input per year (0.304 + 0.002i)
    R_TAX        = 10  # Tax Rate           — input (Year 1 value; others =C10)
    R_MGMT_DRV   = 11  # Annual Mgmt Fee    — input
    R_SPACER     = 12
    R_PL_HDR     = 13  # "PROJECTED P&L" section header
    R_REVENUE    = 14
    R_COGS       = 15
    R_GP         = 16
    R_GM         = 17
    R_EBITDA     = 18
    R_EBITDA_M2  = 19
    R_MGMT_FEE   = 20
    R_ADJ_EBITDA = 21
    R_DA         = 22
    R_EBIT       = 23
    R_INT_EXP    = 24  # hardcoded — from iterative debt engine
    R_EBT        = 25
    R_TAX_CALC   = 26
    R_NI         = 27
    R_NOTE       = 29

    # ── Title ───────────────────────────────────────────────────────────────
    _title_row(ws, R_TITLE, "PROJECTED INCOME STATEMENT  ($M)", n, 2 + n)

    # ── Column headers ───────────────────────────────────────────────────────
    _write(ws, R_HDRS, 2, "Line Item", bold=True, fg=COLOR_SECTION_BG, border=True)
    for yr in range(1, n + 1):
        _col_header(ws, R_HDRS, 2 + yr, f"Year {yr}", bg=COLOR_SECTION_BG)
        ws.cell(row=R_HDRS, column=2 + yr).font = _font(bold=True, color=COLOR_DARK_TEXT)

    # ── Driver block ─────────────────────────────────────────────────────────
    _section_header(ws, R_DRV_HDR, 2,
                    "REVENUE & MARGIN DRIVERS  ← editable inputs",
                    ncols=n, bg=COLOR_NAVY_INPUT)

    def _drv(r, label, values, fmt, bg=COLOR_INPUT_BG):
        _write(ws, r, 2, label, bold=False, fg=bg, border=True)
        for yr, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=2 + yr, value=v)
            cell.font      = _font(bold=True, color=COLOR_DARK_TEXT)
            cell.fill      = _fill(bg)
            cell.alignment = _align(h="right")
            cell.border    = _border()
            cell.number_format = fmt

    # Entry revenue — only Year 1 col has value; rest show note
    _write(ws, R_ENTRY_REV, 2, "Entry LTM Revenue ($M)", bold=False,
           fg=COLOR_INPUT_BG, border=True)
    c = ws.cell(row=R_ENTRY_REV, column=3, value=assumptions.entry_revenue)
    c.font = _font(bold=True, color=COLOR_DARK_TEXT)
    c.fill = _fill(COLOR_INPUT_BG); c.alignment = _align(h="right")
    c.border = _border(); c.number_format = "#,##0.0"
    for yr in range(2, n + 1):
        ws.cell(row=R_ENTRY_REV, column=2 + yr).fill = _fill(COLOR_INPUT_BG)

    _drv(R_GROWTH,     "Revenue Growth (%)",  assumptions.revenue_growth,    "0.0%")
    _drv(R_EBITDA_MRG, "EBITDA Margin (%)",   assumptions.ebitda_margin,     "0.0%",
         bg=COLOR_ROW_ALT)
    _drv(R_DA_PCT,     "D&A (% Revenue)",     assumptions.da_pct_revenue,    "0.0%")
    _drv(R_GM_PCT,     "Gross Margin (%)",    [0.304 + i * 0.002 for i in range(n)],
         "0.0%", bg=COLOR_ROW_ALT)

    # Tax rate — Year 1 is the editable cell; all other years reference it
    _write(ws, R_TAX, 2, "Tax Rate", bold=False, fg=COLOR_INPUT_BG, border=True)
    c = ws.cell(row=R_TAX, column=3, value=assumptions.tax_rate)
    c.font = _font(bold=True, color=COLOR_DARK_TEXT)
    c.fill = _fill(COLOR_INPUT_BG); c.alignment = _align(h="right")
    c.border = _border(); c.number_format = "0.0%"
    for yr in range(2, n + 1):
        c = ws.cell(row=R_TAX, column=2 + yr, value=f"=$C${R_TAX}")
        c.font = _font(color=COLOR_DARK_TEXT); c.fill = _fill(COLOR_INPUT_BG)
        c.alignment = _align(h="right"); c.border = _border()
        c.number_format = "0.0%"

    # Mgmt fee driver — same value every year
    _drv(R_MGMT_DRV, "Annual Mgmt Fee ($M)",
         [assumptions.mgmt_fee_annual] * n, "#,##0.0", bg=COLOR_ROW_ALT)

    # Spacer row
    for c in range(2, 3 + n):
        ws.cell(row=R_SPACER, column=c).fill = _fill(COLOR_LIGHT_GRAY)

    # ── P&L section header ───────────────────────────────────────────────────
    _section_header(ws, R_PL_HDR, 2,
                    "PROJECTED P&L  (formulas reference driver block above)",
                    ncols=n, bg=COLOR_NAVY_MID)

    SECTION_ROWS = {R_REVENUE, R_GP, R_EBITDA, R_EBIT, R_EBT, R_NI}
    BOLD_ROWS    = SECTION_ROWS | {R_ADJ_EBITDA}

    def _label(r, label):
        is_sect = r in SECTION_ROWS
        bg = COLOR_NAVY_MID if is_sect else (COLOR_ROW_ALT if r % 2 == 0 else COLOR_WHITE)
        fc = COLOR_WHITE if is_sect else COLOR_DARK_TEXT
        _write(ws, r, 2, label, bold=(r in BOLD_ROWS),
               fg=bg, font_color=fc, border=True)
        return bg, fc

    def _pf(r, yr, formula, pct=False):
        is_sect = r in SECTION_ROWS
        bg = COLOR_NAVY_MID if is_sect else (COLOR_ROW_ALT if r % 2 == 0 else COLOR_WHITE)
        fc = COLOR_WHITE if is_sect else COLOR_DARK_TEXT
        cell = ws.cell(row=r, column=2 + yr, value=formula)
        cell.font      = _font(bold=(r in BOLD_ROWS), color=fc)
        cell.fill      = _fill(bg)
        cell.alignment = _align(h="right")
        cell.border    = _border()
        cell.number_format = "0.0%" if pct else "#,##0.0"

    # Revenue
    _label(R_REVENUE, "Revenue")
    for yr in range(1, n + 1):
        if yr == 1:
            f = f"=$C${R_ENTRY_REV}*(1+{_yc(1)}{R_GROWTH})"
        else:
            f = f"={_yc(yr-1)}{R_REVENUE}*(1+{_yc(yr)}{R_GROWTH})"
        _pf(R_REVENUE, yr, f)

    # COGS = Revenue × (1 − Gross Margin %)
    _label(R_COGS, "COGS")
    for yr in range(1, n + 1):
        _pf(R_COGS, yr, f"={_yc(yr)}{R_REVENUE}*(1-{_yc(yr)}{R_GM_PCT})")

    # Gross Profit = Revenue − COGS
    _label(R_GP, "Gross Profit")
    for yr in range(1, n + 1):
        _pf(R_GP, yr, f"={_yc(yr)}{R_REVENUE}-{_yc(yr)}{R_COGS}")

    # Gross Margin % = Gross Profit / Revenue
    _label(R_GM, "Gross Margin")
    for yr in range(1, n + 1):
        _pf(R_GM, yr, f"={_yc(yr)}{R_GP}/{_yc(yr)}{R_REVENUE}", pct=True)

    # EBITDA = Revenue × EBITDA Margin
    _label(R_EBITDA, "EBITDA")
    for yr in range(1, n + 1):
        _pf(R_EBITDA, yr, f"={_yc(yr)}{R_REVENUE}*{_yc(yr)}{R_EBITDA_MRG}")

    # EBITDA Margin % = EBITDA / Revenue
    _label(R_EBITDA_M2, "EBITDA Margin")
    for yr in range(1, n + 1):
        _pf(R_EBITDA_M2, yr,
            f"={_yc(yr)}{R_EBITDA}/{_yc(yr)}{R_REVENUE}", pct=True)

    # Mgmt Fee = driver cell
    _label(R_MGMT_FEE, "Mgmt Fee")
    for yr in range(1, n + 1):
        _pf(R_MGMT_FEE, yr, f"={_yc(yr)}{R_MGMT_DRV}")

    # Adj EBITDA = EBITDA + Mgmt Fee
    _label(R_ADJ_EBITDA, "Adj EBITDA")
    for yr in range(1, n + 1):
        _pf(R_ADJ_EBITDA, yr,
            f"={_yc(yr)}{R_EBITDA}+{_yc(yr)}{R_MGMT_FEE}")

    # D&A = Revenue × D&A %
    _label(R_DA, "D&A")
    for yr in range(1, n + 1):
        _pf(R_DA, yr, f"={_yc(yr)}{R_REVENUE}*{_yc(yr)}{R_DA_PCT}")

    # EBIT = EBITDA − D&A
    _label(R_EBIT, "EBIT")
    for yr in range(1, n + 1):
        _pf(R_EBIT, yr, f"={_yc(yr)}{R_EBITDA}-{_yc(yr)}{R_DA}")

    # Interest Expense — hardcoded from iterative debt engine
    _label(R_INT_EXP, "Interest Expense ‡")
    for yr in range(1, n + 1):
        val = float(is_df.loc["Interest Expense", f"Year {yr}"])
        _pf(R_INT_EXP, yr, val)

    # EBT = EBIT − Interest
    _label(R_EBT, "EBT")
    for yr in range(1, n + 1):
        _pf(R_EBT, yr, f"={_yc(yr)}{R_EBIT}-{_yc(yr)}{R_INT_EXP}")

    # Tax = MAX(0, EBT) × Tax Rate  (absolute ref to Year-1 tax cell)
    _label(R_TAX_CALC, "Tax")
    for yr in range(1, n + 1):
        _pf(R_TAX_CALC, yr, f"=MAX(0,{_yc(yr)}{R_EBT})*$C${R_TAX}")

    # Net Income = EBT − Tax
    _label(R_NI, "Net Income")
    for yr in range(1, n + 1):
        _pf(R_NI, yr, f"={_yc(yr)}{R_EBT}-{_yc(yr)}{R_TAX_CALC}")

    # Footnote
    ws.cell(row=R_NOTE, column=2, value=(
        "‡ Interest Expense is computed by an iterative cash-sweep debt model "
        "and cannot be replicated by a simple Excel formula. "
        "All other P&L line items are formula-driven."
    )).font = _font(size=8, italic=True, color="777777")

    # Build row_map for downstream cross-sheet references
    row_map = {lbl: {yr: f"{_yc(yr)}{row}" for yr in range(1, n + 1)}
               for lbl, row in [
                   ("Revenue",       R_REVENUE),
                   ("COGS",          R_COGS),
                   ("Gross Profit",  R_GP),
                   ("Gross Margin",  R_GM),
                   ("EBITDA",        R_EBITDA),
                   ("EBITDA Margin", R_EBITDA_M2),
                   ("D&A",           R_DA),
                   ("EBIT",          R_EBIT),
                   ("Interest Expense", R_INT_EXP),
                   ("EBT",           R_EBT),
                   ("Tax",           R_TAX_CALC),
                   ("Net Income",    R_NI),
               ]}
    return row_map


# ===========================================================================
# Sheet 5 — Cash Flow Statement  (cross-refs IS; subtotals as formulas)
# ===========================================================================

def _build_cash_flow(wb, cfs_df, assumptions, is_row_map=None):
    ws = wb.create_sheet("Cash Flow Statement")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    n = assumptions.hold_years
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    # ── Pre-defined rows ─────────────────────────────────────────────────────
    R_TITLE     = 1
    R_HDRS      = 3
    R_NI        = 4
    R_DA        = 5
    R_DFF       = 6
    R_NWC       = 7
    R_OPCF      = 8
    R_CAPEX     = 9
    R_INVCF     = 10
    R_FCF       = 11
    R_DEBT_REP  = 12
    R_MGMT_FEE  = 13
    R_FINCF     = 14
    R_NET_CHG   = 15

    _title_row(ws, R_TITLE, "PROJECTED CASH FLOW STATEMENT  ($M)", n, 2 + n)

    _write(ws, R_HDRS, 2, "Line Item", bold=True, fg=COLOR_SECTION_BG, border=True)
    for yr in range(1, n + 1):
        _col_header(ws, R_HDRS, 2 + yr, f"Year {yr}", bg=COLOR_SECTION_BG)
        ws.cell(row=R_HDRS, column=2 + yr).font = _font(bold=True, color=COLOR_DARK_TEXT)

    SECTION_ROWS = {R_OPCF, R_FCF, R_NET_CHG}
    BOLD_ROWS    = SECTION_ROWS

    def _label(r, label):
        is_sect = r in SECTION_ROWS
        bg = COLOR_NAVY_MID if is_sect else (COLOR_ROW_ALT if r % 2 == 0 else COLOR_WHITE)
        fc = COLOR_WHITE if is_sect else COLOR_DARK_TEXT
        _write(ws, r, 2, label, bold=(r in BOLD_ROWS), fg=bg, font_color=fc, border=True)
        return bg, fc

    def _cf(r, yr, val_or_f, note=""):
        is_sect = r in SECTION_ROWS
        bg = COLOR_NAVY_MID if is_sect else (COLOR_ROW_ALT if r % 2 == 0 else COLOR_WHITE)
        fc = COLOR_WHITE if is_sect else COLOR_DARK_TEXT
        cell = ws.cell(row=r, column=2 + yr, value=val_or_f)
        cell.font = _font(bold=(r in BOLD_ROWS), color=fc)
        cell.fill = _fill(bg); cell.alignment = _align(h="right")
        cell.border = _border(); cell.number_format = "#,##0.0"

    IS = "Income Statement"

    # Net Income — cross-ref to IS
    _label(R_NI, "Net Income")
    for yr in range(1, n + 1):
        if is_row_map and "Net Income" in is_row_map:
            f = f"='{IS}'!{is_row_map['Net Income'][yr]}"
        else:
            f = float(cfs_df.loc["Net Income", f"Year {yr}"])
        _cf(R_NI, yr, f)

    # (+) D&A — cross-ref to IS
    _label(R_DA, "(+) D&A")
    for yr in range(1, n + 1):
        if is_row_map and "D&A" in is_row_map:
            f = f"='{IS}'!{is_row_map['D&A'][yr]}"
        else:
            f = float(cfs_df.loc["(+) D&A", f"Year {yr}"])
        _cf(R_DA, yr, f)

    # (+) DFF Amort — hardcoded (from debt engine)
    _label(R_DFF, "(+) DFF Amort")
    for yr in range(1, n + 1):
        _cf(R_DFF, yr, float(cfs_df.loc["(+) DFF Amort", f"Year {yr}"]))

    # Δ NWC — hardcoded (working capital calc)
    _label(R_NWC, "Δ NWC")
    for yr in range(1, n + 1):
        _cf(R_NWC, yr, float(cfs_df.loc["Δ NWC", f"Year {yr}"]))

    # Operating CF = SUM(NI:NWC)
    _label(R_OPCF, "Operating CF")
    for yr in range(1, n + 1):
        f = (f"={_yc(yr)}{R_NI}+{_yc(yr)}{R_DA}"
             f"+{_yc(yr)}{R_DFF}+{_yc(yr)}{R_NWC}")
        _cf(R_OPCF, yr, f)

    # (-) CapEx = − Revenue × CapEx %  (cross-ref IS revenue if available)
    _label(R_CAPEX, "(-) CapEx")
    for yr in range(1, n + 1):
        cpct = assumptions.capex_pct_revenue[yr - 1]
        if is_row_map and "Revenue" in is_row_map:
            f = f"=-'{IS}'!{is_row_map['Revenue'][yr]}*{cpct}"
        else:
            f = float(cfs_df.loc["(-) CapEx", f"Year {yr}"])
        _cf(R_CAPEX, yr, f)

    # Investing CF = CapEx (same cell)
    _label(R_INVCF, "Investing CF")
    for yr in range(1, n + 1):
        _cf(R_INVCF, yr, f"={_yc(yr)}{R_CAPEX}")

    # Free Cash Flow = Operating CF + Investing CF
    _label(R_FCF, "Free Cash Flow")
    for yr in range(1, n + 1):
        _cf(R_FCF, yr, f"={_yc(yr)}{R_OPCF}+{_yc(yr)}{R_INVCF}")

    # (-) Debt Repaid — hardcoded (cash sweep model)
    _label(R_DEBT_REP, "(-) Debt Repaid")
    for yr in range(1, n + 1):
        _cf(R_DEBT_REP, yr, float(cfs_df.loc["(-) Debt Repaid", f"Year {yr}"]))

    # (-) Mgmt Fee — hardcoded
    _label(R_MGMT_FEE, "(-) Mgmt Fee")
    for yr in range(1, n + 1):
        _cf(R_MGMT_FEE, yr, float(cfs_df.loc["(-) Mgmt Fee", f"Year {yr}"]))

    # Financing CF = Debt Repaid + Mgmt Fee (both stored as negative)
    _label(R_FINCF, "Financing CF")
    for yr in range(1, n + 1):
        _cf(R_FINCF, yr, f"={_yc(yr)}{R_DEBT_REP}+{_yc(yr)}{R_MGMT_FEE}")

    # Net Change in Cash = OpCF + InvCF + FinCF
    _label(R_NET_CHG, "Net Change in Cash")
    for yr in range(1, n + 1):
        _cf(R_NET_CHG, yr,
            f"={_yc(yr)}{R_OPCF}+{_yc(yr)}{R_INVCF}+{_yc(yr)}{R_FINCF}")


# ===========================================================================
# Sheet 6 — Balance Sheet  (SUM/formula totals and BS Check)
# ===========================================================================

def _build_balance_sheet(wb, opening_bs, bs_df, assumptions):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 18
    n = assumptions.hold_years
    for i in range(n):
        ws.column_dimensions[get_column_letter(4 + i)].width = 16

    _title_row(ws, 1, "BALANCE SHEET  ($M)", n, 3 + n)

    r = 3
    _write(ws, r, 2, "Line Item", bold=True, fg=COLOR_SECTION_BG, border=True)
    _col_header(ws, r, 3, "Day 1 (Opening)", bg=COLOR_SECTION_BG)
    ws.cell(row=r, column=3).font = _font(bold=True, color=COLOR_DARK_TEXT)
    for i in range(n):
        _col_header(ws, r, 4 + i, f"Year {i+1}", bg=COLOR_SECTION_BG)
        ws.cell(row=r, column=4 + i).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    # Map item name → row number for formula references
    item_row = {}

    FORMULA_ROWS = {"Total Current Assets", "Total Assets",
                    "Total Current Liabilities", "Total Liabilities",
                    "Total Equity", "Total L+E", "BS Check (Assets - L+E)"}
    BOLD_ROWS    = {"Total Current Assets", "Total Assets", "Total Current Liabilities",
                    "Total Liabilities", "Total Equity", "Total L+E"}
    SECTION_ROWS = {"Total Assets", "Total L+E"}
    CHECK_ROWS   = {"BS Check (Assets - L+E)"}

    # Track row ranges for SUM formulas
    ca_start = ca_end = None      # Current Asset items (not Total CA)
    cl_start = cl_end = None      # Current Liability items (not Total CL)
    nca_rows: list[int] = []      # Non-current asset item rows
    ncl_rows: list[int] = []      # Non-current liability item rows

    for label in opening_bs.index:
        item_row[label] = r
        is_sect  = label in SECTION_ROWS
        is_bold  = label in BOLD_ROWS
        is_check = label in CHECK_ROWS
        is_form  = label in FORMULA_ROWS

        if is_check:
            bg, fc = "FFF0F0", "8B0000"
        elif is_sect:
            bg, fc = COLOR_NAVY_MID, COLOR_WHITE
        else:
            bg, fc = (COLOR_ROW_ALT if r % 2 else COLOR_WHITE), COLOR_DARK_TEXT

        _write(ws, r, 2, label, bold=is_bold, fg=bg, font_color=fc, border=True)

        # Track ranges for SUM formulas
        if label == "Cash & Equivalents":        ca_start = r
        if label == "Other Current Assets":      ca_end   = r
        if label == "PP&E (net)":                nca_rows.append(r)
        if label in ("Operating Lease ROU Assets", "Goodwill",
                     "Deferred Financing Fees", "Other LT Assets"):
            nca_rows.append(r)
        if label == "Accounts Payable":          cl_start = r
        if label == "Other Current Liabilities": cl_end   = r
        # Opening BS uses "Total Debt (new structure)"; projected BS uses "Total Debt"
        if label in ("Total Debt (new structure)", "Total Debt",
                     "LT Operating Lease", "Other LT Liabilities"):
            ncl_rows.append(r)

        # Opening column (col 3) — always values
        val_open = float(opening_bs[label])
        open_cell = ws.cell(row=r, column=3, value=val_open)
        open_cell.font = _font(bold=is_bold, color=fc)
        open_cell.fill = _fill(bg); open_cell.alignment = _align(h="right")
        open_cell.border = _border(); open_cell.number_format = "#,##0.0"

        # Alias: opening BS label → projected BS label (where they differ)
        BS_LABEL_ALIAS = {"Total Debt (new structure)": "Total Debt"}

        # Projected columns (cols 4..3+n)
        for j in range(n):
            col = 4 + j
            col_name = f"Year {j+1}"
            yr = j + 1

            proj_label = BS_LABEL_ALIAS.get(label, label)
            if proj_label in bs_df.index:
                raw = float(bs_df.loc[proj_label, col_name])
            else:
                raw = 0.0

            # Write formula for total/check rows; value for everything else
            if label == "Total Current Assets" and ca_start and ca_end:
                val = f"=SUM({get_column_letter(col)}{ca_start}:{get_column_letter(col)}{ca_end})"
            elif label == "Total Assets" and nca_rows:
                tca_r = item_row.get("Total Current Assets", r)
                parts = "+".join(f"{get_column_letter(col)}{nr}" for nr in nca_rows)
                val = f"={get_column_letter(col)}{tca_r}+{parts}"
            elif label == "Total Current Liabilities" and cl_start and cl_end:
                val = f"=SUM({get_column_letter(col)}{cl_start}:{get_column_letter(col)}{cl_end})"
            elif label == "Total Liabilities" and ncl_rows:
                tcl_r = item_row.get("Total Current Liabilities", r)
                parts = "+".join(f"{get_column_letter(col)}{nr}" for nr in ncl_rows)
                val = f"={get_column_letter(col)}{tcl_r}+{parts}"
            elif label == "Total Equity":
                paid_r = item_row.get("Sponsor Equity (Paid-In)", r)
                re_r   = item_row.get("Retained Earnings", r)
                val = f"={get_column_letter(col)}{paid_r}+{get_column_letter(col)}{re_r}"
            elif label == "Total L+E":
                tl_r = item_row.get("Total Liabilities", r)
                te_r = item_row.get("Total Equity", r)
                val = f"={get_column_letter(col)}{tl_r}+{get_column_letter(col)}{te_r}"
            elif label == "BS Check (Assets - L+E)":
                ta_r  = item_row.get("Total Assets", r)
                tle_r = item_row.get("Total L+E", r)
                val = f"={get_column_letter(col)}{ta_r}-{get_column_letter(col)}{tle_r}"
            else:
                val = raw

            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _font(bold=is_bold, color=fc)
            cell.fill = _fill(bg); cell.alignment = _align(h="right")
            cell.border = _border(); cell.number_format = "#,##0.0"

        r += 1


# ===========================================================================
# Sheet 7 — Debt Schedule  (Ending Balance as formula)
# ===========================================================================

def _build_debt_schedule(wb, debt_sched, assumptions):
    ws = wb.create_sheet("Debt Schedule")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    n = assumptions.hold_years
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    _title_row(ws, 1, "DEBT SCHEDULE & CASH SWEEP  ($M)", n, 2 + n)

    summary_df  = debt_sched["summary_df"]
    tranche_dfs = debt_sched["tranche_dfs"]

    r = 3
    _section_header(ws, r, 2, "CONSOLIDATED DEBT SUMMARY", ncols=n)
    r += 1

    _write(ws, r, 2, "Metric", bold=True, fg=COLOR_SECTION_BG, border=True)
    for i in range(n):
        _col_header(ws, r, 3 + i, f"Year {i+1}", bg=COLOR_SECTION_BG)
        ws.cell(row=r, column=3 + i).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    BOLD_ROWS_DS = {"End. Total Debt ($M)", "Interest Expense ($M)", "Gross Leverage (x)"}
    # Track rows for formula: Ending = Beginning − Req Amort − Cash Sweep
    beg_row = req_row = sweep_row = end_row = None

    metric_cols = [c for c in summary_df.columns if c != "Year"]
    for i, col in enumerate(metric_cols):
        is_bold = col in BOLD_ROWS_DS
        bg      = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        is_lev  = "(x)" in col
        _write(ws, r, 2, col, bold=is_bold, fg=bg, border=True)

        # Track key rows
        if col == "Beg. Total Debt ($M)":  beg_row  = r
        if col == "Req. Amortization ($M)": req_row  = r
        if col == "Cash Sweep ($M)":        sweep_row = r

        for j in range(n):
            val = summary_df.iloc[j][col] if j < len(summary_df) else None
            try:
                fval = float(val) if val is not None else None
            except (TypeError, ValueError):
                fval = None

            # Ending debt = Beginning − Req Amort − Cash Sweep (formula)
            if col == "End. Total Debt ($M)" and beg_row and req_row and sweep_row:
                c_l = get_column_letter(3 + j)
                val_or_f = (f"={c_l}{beg_row}"
                            f"-{c_l}{req_row}"
                            f"-{c_l}{sweep_row}")
                end_row = r
            else:
                val_or_f = fval

            cell = ws.cell(row=r, column=3 + j, value=val_or_f)
            cell.font = _font(bold=is_bold)
            cell.fill = _fill(bg); cell.alignment = _align(h="right")
            cell.border = _border()
            cell.number_format = '0.00"x"' if is_lev else "#,##0.0"
        r += 1

    r += 1

    # Per-tranche detail
    for tranche_name, df in tranche_dfs.items():
        _section_header(ws, r, 2, f"TRANCHE: {tranche_name.upper()}", ncols=n)
        r += 1
        col_labels = [c for c in df.columns if c != "Year"]
        _write(ws, r, 2, "Item", bold=True, fg=COLOR_SECTION_BG, border=True)
        for i in range(n):
            _col_header(ws, r, 3 + i, f"Year {i+1}", bg=COLOR_SECTION_BG)
            ws.cell(row=r, column=3 + i).font = _font(bold=True, color=COLOR_DARK_TEXT)
        r += 1

        t_beg_row = t_req_row = t_sweep_row = None
        for i, lbl in enumerate(col_labels):
            bg     = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
            is_rate = lbl == "Rate"
            _write(ws, r, 2, lbl, bold=False, fg=bg, border=True)
            if lbl == "Beginning Balance ($M)": t_beg_row   = r
            if lbl == "Required Amort ($M)":    t_req_row   = r
            if lbl == "Cash Sweep ($M)":        t_sweep_row = r

            for j in range(n):
                raw = df.iloc[j][lbl] if j < len(df) else None
                if is_rate:
                    cell = ws.cell(row=r, column=3 + j, value=raw)
                    cell.number_format = "@"
                elif (lbl == "Ending Balance ($M)"
                      and t_beg_row and t_req_row and t_sweep_row):
                    c_l = get_column_letter(3 + j)
                    cell = ws.cell(row=r, column=3 + j,
                                   value=f"={c_l}{t_beg_row}-{c_l}{t_req_row}-{c_l}{t_sweep_row}")
                    cell.number_format = "#,##0.0"
                else:
                    try:
                        fval = float(raw) if raw is not None else None
                    except (TypeError, ValueError):
                        fval = None
                    cell = ws.cell(row=r, column=3 + j, value=fval)
                    cell.number_format = "#,##0.0"
                cell.font = _font(); cell.fill = _fill(bg)
                cell.alignment = _align(h="right"); cell.border = _border()
            r += 1
        r += 1


# ===========================================================================
# Sheet 8 — Returns & Value Bridge  (bridge total and exit equity as formulas)
# ===========================================================================

def _build_returns(wb, returns, assumptions):
    ws = wb.create_sheet("Returns & Value Bridge")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    r_data = returns
    a = assumptions

    _title_row(ws, 1, "RETURNS ANALYSIS & VALUE ATTRIBUTION  ($M)", 1, 5)

    r = 3
    _section_header(ws, r, 2, "RETURNS SUMMARY", ncols=3)
    r += 1

    irr  = r_data["irr"]
    moic = r_data["moic"]

    returns_items = [
        ("Equity Invested ($M)",   r_data["entry_equity"],   '#,##0.0', False),
        ("Exit Equity ($M)",       r_data["exit_equity"],    '#,##0.0', False),
        ("IRR",                    irr if not np.isnan(irr) else None, "0.0%", True),
        ("MOIC",                   moic,                     '0.00"x"', True),
        ("Hold Period (years)",    r_data["hold_years"],     '0',        False),
        ("Entry EV ($M)",          r_data["entry_ev"],       '#,##0.0', False),
        ("Entry EV/EBITDA (x)",    r_data["entry_ev_multiple"], '0.0"x"', False),
        ("Entry EBITDA ($M)",      r_data["entry_ebitda"],   '#,##0.0', False),
        ("Entry Net Debt ($M)",    r_data["entry_net_debt"], '#,##0.0', False),
        ("Exit EV ($M)",           r_data["exit_ev"],        '#,##0.0', False),
        ("Exit EV/EBITDA (x)",     r_data["exit_ev_multiple"], '0.0"x"', False),
        ("Exit EBITDA ($M)",       r_data["exit_ebitda"],    '#,##0.0', False),
        ("Exit Net Debt ($M)",     r_data["exit_net_debt"],  '#,##0.0', False),
    ]

    for i, (lbl, val, fmt, highlight) in enumerate(returns_items):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        fc = COLOR_GREEN if highlight else COLOR_DARK_TEXT
        _write(ws, r, 2, lbl, bold=highlight, fg=bg, font_color=COLOR_DARK_TEXT, border=True)
        cell = ws.cell(row=r, column=3, value=val)
        cell.font = _font(bold=highlight, color=fc)
        cell.fill = _fill(bg); cell.alignment = _align(h="right")
        cell.border = _border(); cell.number_format = fmt
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=4).fill = _fill(bg)
        r += 1

    r += 1
    _section_header(ws, r, 2, "VALUE CREATION BRIDGE  ($M)", ncols=3)
    r += 1

    bridge = r_data["bridge"]
    bridge_order = [
        ("Entry Equity Value ($M)",                   "Entry Equity Value ($M)"),
        ("  (+) EBITDA Growth",                       "EBITDA Growth ($M)"),
        ("  (+) Multiple Expansion / (Contraction)",  "Multiple Expansion / (Contraction) ($M)"),
        ("  (+) Debt Paydown (Deleveraging)",          "Deleveraging ($M)"),
        ("Total Value Creation ($M)",                  None),  # formula
        ("Exit Equity Value ($M)",                     None),  # formula
    ]

    comp_rows = {}
    for i, (lbl, key) in enumerate(bridge_order):
        is_total = "Total" in lbl or "Entry" in lbl or "Exit" in lbl
        bg = COLOR_SECTION_BG if is_total else (COLOR_ROW_ALT if i % 2 else COLOR_WHITE)
        _write(ws, r, 2, lbl, bold=is_total, fg=bg, border=True)
        comp_rows[lbl] = r

        if lbl == "Total Value Creation ($M)":
            ebitda_r   = comp_rows.get("  (+) EBITDA Growth", r)
            mult_r     = comp_rows.get("  (+) Multiple Expansion / (Contraction)", r)
            delev_r    = comp_rows.get("  (+) Debt Paydown (Deleveraging)", r)
            val = f"=C{ebitda_r}+C{mult_r}+C{delev_r}"
        elif lbl == "Exit Equity Value ($M)":
            entry_r = comp_rows.get("Entry Equity Value ($M)", r)
            total_r = comp_rows.get("Total Value Creation ($M)", r)
            val = f"=C{entry_r}+C{total_r}"
        else:
            val = float(bridge.get(key, 0))

        cell = ws.cell(row=r, column=3, value=val)
        cell.font = _font(bold=is_total); cell.fill = _fill(bg)
        cell.alignment = _align(h="right"); cell.border = _border()
        cell.number_format = "#,##0.0"
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=4).fill = _fill(bg)
        r += 1

    r += 1
    _section_header(ws, r, 2, "CASH FLOW TO EQUITY  ($M)", ncols=3)
    r += 1

    n = a.hold_years
    _write(ws, r, 2, "Year", bold=True, fg=COLOR_SECTION_BG, border=True)
    _col_header(ws, r, 3, "Cash Flow ($M)", bg=COLOR_SECTION_BG)
    ws.cell(row=r, column=3).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    for yr in range(n + 1):
        bg = COLOR_ROW_ALT if yr % 2 else COLOR_WHITE
        label = f"Year {yr} (Equity In)" if yr == 0 else (
                f"Year {yr} (Exit)" if yr == n else f"Year {yr}")
        is_bold = yr in {0, n}
        if yr == n:
            bg = COLOR_SECTION_BG
        _write(ws, r, 2, label, bold=is_bold, fg=bg, border=True)
        val = (-r_data["entry_equity"] if yr == 0
               else r_data["exit_equity"] if yr == n
               else 0.0)
        cell = ws.cell(row=r, column=3, value=float(val))
        cell.font = _font(bold=is_bold); cell.fill = _fill(bg)
        cell.alignment = _align(h="right"); cell.border = _border()
        cell.number_format = "#,##0.0"
        r += 1


# ===========================================================================
# Sheet 9 — Credit Metrics
# ===========================================================================

def _build_credit_metrics(wb, credit_extra, assumptions):
    ws = wb.create_sheet("Credit Metrics")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    n = assumptions.hold_years
    for i in range(n):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    credit_df   = credit_extra["credit_df"]
    covenant_df = credit_extra["covenant_df"]

    _title_row(ws, 1, "CREDIT METRICS & COVENANT ANALYSIS", n, 2 + n)

    r = 3
    _section_header(ws, r, 2, "YEAR-BY-YEAR CREDIT METRICS", ncols=n)
    r += 1

    _write(ws, r, 2, "Metric", bold=True, fg=COLOR_SECTION_BG, border=True)
    for i in range(n):
        _col_header(ws, r, 3 + i, f"Year {i+1}", bg=COLOR_SECTION_BG)
        ws.cell(row=r, column=3 + i).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    PCT_METRICS  = {"EBITDA Margin", "FCF / EBITDA", "CapEx / Revenue"}
    MULT_METRICS = {"Gross Leverage (x)", "Net Leverage (x)", "Interest Coverage (x)",
                    "Fixed Charge Coverage (x)", "DSCR (x)"}
    BOLD_METRICS = {"Gross Leverage (x)", "Net Leverage (x)", "Free Cash Flow ($M)"}

    for i, col in enumerate([c for c in credit_df.columns if c != "Year"]):
        is_pct  = col in PCT_METRICS
        is_mult = col in MULT_METRICS
        is_bold = col in BOLD_METRICS
        is_str  = col == "Implied Rating"
        bg      = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, r, 2, col, bold=is_bold, fg=bg, border=True)
        for j in range(n):
            val = credit_df.iloc[j][col] if j < len(credit_df) else None
            if is_str:
                cell = ws.cell(row=r, column=3 + j, value=str(val) if val else None)
                cell.number_format = "@"
            else:
                try:
                    fval = float(val) if val is not None else None
                except (TypeError, ValueError):
                    fval = None
                cell = ws.cell(row=r, column=3 + j, value=fval)
                cell.number_format = ("0.0%" if is_pct else
                                      '0.00"x"' if is_mult else "#,##0.0")
            cell.font = _font(bold=is_bold); cell.fill = _fill(bg)
            cell.alignment = _align(h="right"); cell.border = _border()
        r += 1

    r += 1
    _section_header(ws, r, 2, "COVENANT COMPLIANCE & HEADROOM", ncols=n)
    r += 1
    _write(ws, r, 2, "Covenant Metric", bold=True, fg=COLOR_SECTION_BG, border=True)
    for i in range(n):
        _col_header(ws, r, 3 + i, f"Year {i+1}", bg=COLOR_SECTION_BG)
        ws.cell(row=r, column=3 + i).font = _font(bold=True, color=COLOR_DARK_TEXT)
    r += 1

    for i, col in enumerate([c for c in covenant_df.columns if c != "Year"]):
        bg  = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        is_compliance = "In Compliance" in col
        _write(ws, r, 2, col, bold=is_compliance, fg=bg, border=True)
        for j in range(n):
            val = covenant_df.iloc[j][col] if j < len(covenant_df) else None
            if is_compliance:
                cell = ws.cell(row=r, column=3 + j, value=str(val) if val else None)
                cell.number_format = "@"
                cell.font = _font(bold=True,
                                  color=COLOR_GREEN if str(val) == "YES" else COLOR_RED_DARK)
            else:
                try:
                    fval = float(val) if val is not None else None
                except (TypeError, ValueError):
                    fval = None
                cell = ws.cell(row=r, column=3 + j, value=fval)
                cell.font = _font()
                cell.number_format = '0.00"x"' if fval is not None else "@"
            cell.fill = _fill(bg); cell.alignment = _align(h="right")
            cell.border = _border()
        r += 1


# ===========================================================================
# Sheet 10 — Scenario Analysis
# ===========================================================================

def _build_scenarios(wb, assumptions):
    ws = wb.create_sheet("Scenario Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    scenario_result = run_scenarios(base_assumptions=assumptions)
    comp_df    = scenario_result["comparison_df"]
    returns_df = scenario_result["returns_df"]

    _title_row(ws, 1, "SCENARIO ANALYSIS — BULL / BASE / BEAR", 1, 5)

    r = 3
    _section_header(ws, r, 2, "RETURNS COMPARISON", ncols=4)
    r += 1

    _write(ws, r, 2, "Metric", bold=True, fg=COLOR_SECTION_BG, border=True)
    for i, sc in enumerate(["Bear", "Base", "Bull"]):
        sc_bg = {"Bear": COLOR_RED_DARK, "Base": COLOR_NAVY_MID, "Bull": COLOR_GREEN}[sc]
        _col_header(ws, r, 3 + i, sc, bg=sc_bg)
    r += 1

    for i, idx_label in enumerate(returns_df.index):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        is_bold = idx_label in {"IRR", "MOIC"}
        _write(ws, r, 2, idx_label, bold=is_bold, fg=bg, border=True)
        for j, sc in enumerate(["Bear", "Base", "Bull"]):
            val = returns_df.loc[idx_label, sc]
            cell = ws.cell(row=r, column=3 + j, value=val)
            cell.font = _font(bold=is_bold); cell.fill = _fill(bg)
            cell.alignment = _align(h="right"); cell.border = _border()
            cell.number_format = "@"
        r += 1

    r += 1
    _section_header(ws, r, 2, "DETAILED COMPARISON", ncols=4)
    r += 1
    _write(ws, r, 2, "Metric", bold=True, fg=COLOR_SECTION_BG, border=True)
    for i, sc in enumerate(["Bear", "Base", "Bull"]):
        sc_bg = {"Bear": COLOR_RED_DARK, "Base": COLOR_NAVY_MID, "Bull": COLOR_GREEN}[sc]
        _col_header(ws, r, 3 + i, sc, bg=sc_bg)
    r += 1
    for i, idx_label in enumerate(comp_df.index):
        bg = COLOR_ROW_ALT if i % 2 else COLOR_WHITE
        _write(ws, r, 2, idx_label, bold=False, fg=bg, border=True)
        for j, sc in enumerate(comp_df.columns):
            cell = ws.cell(row=r, column=3 + j, value=comp_df.loc[idx_label, sc])
            cell.font = _font(); cell.fill = _fill(bg)
            cell.alignment = _align(h="right"); cell.border = _border()
            cell.number_format = "@"
        r += 1


# ===========================================================================
# Master builder
# ===========================================================================

def build_excel_workbook(assumptions: DealAssumptions | None = None) -> bytes:
    """
    Build the complete LBO workbook and return it as bytes.

    Parameters
    ----------
    assumptions : DealAssumptions (default: base_case())

    Returns
    -------
    bytes  — pass directly to st.download_button or write to disk
    """
    if assumptions is None:
        assumptions = base_case()

    result       = run_model(assumptions)
    credit_extra = build_credit_dashboard(result, assumptions)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_cover(wb, assumptions, result)
    _build_assumptions_sheet(wb, assumptions)
    _build_sources_uses(wb, assumptions)

    # IS returns row_map so CFS can cross-reference it
    is_row_map = _build_income_statement(wb, result["is_df"], assumptions)

    _build_cash_flow(wb, result["cfs_df"], assumptions, is_row_map=is_row_map)
    _build_balance_sheet(wb, result["opening_bs"], result["bs_df"], assumptions)
    _build_debt_schedule(wb, result["debt_schedule"], assumptions)
    _build_returns(wb, result["returns"], assumptions)
    _build_credit_metrics(wb, credit_extra, assumptions)
    _build_scenarios(wb, assumptions)

    wb.active = wb["Cover"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Building formula-driven LBO workbook…")
    data = build_excel_workbook()
    fname = "DollarGeneral_LBO_Model.xlsx"
    with open(fname, "wb") as f:
        f.write(data)
    print(f"Saved: {fname}  ({len(data):,} bytes)")
